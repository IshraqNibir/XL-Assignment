from openpyxl import load_workbook
workbook = load_workbook(filename="assignment.xlsx")
workbook.sheetnames
sheet = workbook.active
origin_cnt = {}
origin_wise_total_acc = {} 
most_frequent = {}
least_frequent = {}
for value in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
    if value[8] in origin_cnt:
        if value[7] > most_frequent[value[8]]:
            most_frequent[value[8]] = value[7]
        if value[7] < least_frequent[value[8]]:
            least_frequent[value[8]] = value[7]

        origin_cnt[value[8]] = origin_cnt[value[8]] + 1
        origin_wise_total_acc[value[8]] = origin_wise_total_acc[value[8]] + value[6]
    else:
        origin_cnt[value[8]] = 1
        origin_wise_total_acc[value[8]] = value[6]
        most_frequent[value[8]] = value[7]
        least_frequent[value[8]] = value[7]
# print(origin_cnt, origin_wise_total_acc)

print()
print("Total Acceleration Across Origin:")
for k,v in origin_wise_total_acc.items():
    print(k, v)

print()
print("Average Acceleration Across Origin:")
for k,v in origin_wise_total_acc.items():
    print(k, v/origin_cnt[k])

print()
print("Most Frequent Model For Origin")
for k,v in most_frequent.items():
    print(k, v)

print()
print("Least Frequent Model For Origin")
for k,v in least_frequent.items():
    print(k, v)







# print(sheet.cell(row=1, column=5).value)